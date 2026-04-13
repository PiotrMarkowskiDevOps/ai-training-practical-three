# scheduler.py
#
# Build your bootcamp scheduler here.
#
# Work through the challenges in order:
#   Run /next in Claude Code to see what to do next.
#   Run /spec to write a specification for the next function.
#   Run /build to implement from your spec.
#   Run python3 -m pytest tests/ -v to verify what you've built.
#
# Each challenge adds new functions — don't remove old ones.

from collections import defaultdict
from datetime import date
import holidays as hols
import openpyxl
from ortools.sat.python import cp_model

LOCATION_COUNTRY_CODES: dict[str, str] = {
    "london": "GB",
    "manchester": "GB",
    "bristol": "GB",
    "paris": "FR",
    "amsterdam": "NL",
    "stockholm": "SE",
}


def parse_availability(filepath: str) -> tuple:
    """Read trainer availability from an xlsx file.

    Basic format (basic.xlsx): returns (dates, trainers).
    Extended format (intermediate/advanced.xlsx): returns (dates, trainers, trainer_info, weekly_caps).
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = rows[0]

    # Detect format: try parsing header[1] as a date
    extended = False
    try:
        date.fromisoformat(str(header[1]))
    except (ValueError, TypeError):
        extended = True

    date_start = 4 if extended else 1

    # Parse dates from header (starting at date_start)
    dates: list[date] = []
    for cell in header[date_start:]:
        if cell is not None:
            try:
                dates.append(date.fromisoformat(str(cell)))
            except (ValueError, TypeError):
                pass
    dates = sorted(dates)

    trainers: dict[str, dict[date, bool]] = {}
    trainer_info: dict[str, dict] = {}
    weekly_caps: dict[tuple[int, int], int] = {}
    cap_row = None

    for row in rows[1:]:
        if not row:
            continue
        if row[0] is None:
            if extended:
                cap_row = row
            continue
        name = str(row[0]).strip()
        avail: dict[date, bool] = {}
        for i, d in enumerate(dates):
            col_idx = date_start + i
            cell_val = row[col_idx] if col_idx < len(row) else None
            avail[d] = cell_val == "Yes"
        trainers[name] = avail
        if extended:
            trainer_info[name] = {
                "home_location": str(row[1]).strip() if row[1] else None,
                "french_speaking": str(row[2]).strip() == "Yes" if row[2] else False,
                "max_per_week": int(row[3]) if row[3] is not None else None,
            }

    # Parse weekly caps from cap row (values appear on Monday columns only)
    if extended and cap_row is not None:
        for i, d in enumerate(dates):
            col_idx = date_start + i
            cap_val = cap_row[col_idx] if col_idx < len(cap_row) else None
            if cap_val is not None:
                yr, wk, _ = d.isocalendar()
                weekly_caps[(yr, wk)] = int(cap_val)

    if extended:
        return (dates, trainers, trainer_info, weekly_caps)
    return (dates, trainers)


def generate_slots(
    dates: list[date],
    trainers: dict[str, dict[date, bool]],
    pattern: str = "mon-tue,thu-fri",
) -> list[tuple[date, date]]:
    """Find all valid bootcamp slots from available dates and trainer availability.

    A slot is a (day1, day2) pair matching the pattern (Mon-Tue or Thu-Fri)
    where at least 2 trainers are available on both days.
    """
    patterns = [p.strip().lower() for p in pattern.split(",")]
    allowed_weekday_pairs: set[tuple[int, int]] = set()
    if "mon-tue" in patterns:
        allowed_weekday_pairs.add((0, 1))
    if "thu-fri" in patterns:
        allowed_weekday_pairs.add((3, 4))

    slots: list[tuple[date, date]] = []
    for i in range(len(dates) - 1):
        d1, d2 = dates[i], dates[i + 1]
        if (d1.weekday(), d2.weekday()) not in allowed_weekday_pairs:
            continue
        available_both = [
            name for name, avail in trainers.items()
            if avail.get(d1, False) and avail.get(d2, False)
        ]
        if len(available_both) >= 2:
            slots.append((d1, d2))

    return slots


def schedule_greedy(
    dates: list[date],
    trainers: dict[str, dict[date, bool]],
    slots: list[tuple[date, date]],
    config: dict | None = None,
) -> list[dict]:
    """Greedily assign 2 trainers to each slot, skipping slots where fewer than 2 are free.

    Trainers are selected alphabetically for determinism. A trainer can only
    be booked once per day across all scheduled bootcamps.
    """
    booked: set[tuple[str, date]] = set()
    schedule: list[dict] = []
    experienced: set[str] = config.get("experienced", set()) if config else set()

    for d1, d2 in slots:
        candidates = sorted(
            name for name, avail in trainers.items()
            if avail.get(d1, False)
            and avail.get(d2, False)
            and (name, d1) not in booked
            and (name, d2) not in booked
        )
        if len(candidates) < 2:
            continue

        if experienced:
            exp_candidates = [c for c in candidates if c in experienced]
            if not exp_candidates:
                continue
            first = exp_candidates[0]
            remaining = [c for c in candidates if c != first]
            if not remaining:
                continue
            assigned = sorted([first, remaining[0]])
        else:
            assigned = candidates[:2]

        for name in assigned:
            booked.add((name, d1))
            booked.add((name, d2))
        schedule.append({"slot": (d1, d2), "trainers": assigned})

    return schedule


def schedule_optimal(
    dates: list[date],
    trainers: dict[str, dict[date, bool]],
    slots: list[tuple[date, date]],
    config: dict | None = None,
    trainer_info: dict | None = None,
    weekly_caps: dict | None = None,
    locations: list[dict] | None = None,
    weightings: dict | None = None,
) -> list[dict]:
    """CP-SAT solver that maximises total bootcamps subject to all constraints."""
    if not slots:
        return []

    model = cp_model.CpModel()
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 30

    trainer_names = sorted(trainers.keys())
    n_trainers = len(trainer_names)
    n_slots = len(slots)
    experienced: set[str] = config.get("experienced", set()) if config else set()

    # x[t_idx, s_idx]: trainer t assigned to slot s — only when trainer available both days
    x: dict[tuple[int, int], cp_model.IntVar] = {}
    for t_idx, name in enumerate(trainer_names):
        for s_idx, (d1, d2) in enumerate(slots):
            if trainers[name].get(d1, False) and trainers[name].get(d2, False):
                x[t_idx, s_idx] = model.new_bool_var(f"x_{t_idx}_{s_idx}")

    # active[s]: slot s is scheduled
    active = [model.new_bool_var(f"active_{s}") for s in range(n_slots)]

    # Exactly 2 trainers per active slot
    for s_idx in range(n_slots):
        slot_vars = [x[t_idx, s_idx] for t_idx in range(n_trainers) if (t_idx, s_idx) in x]
        model.add(sum(slot_vars) == 2 * active[s_idx])

    # No double-booking: each trainer at most 1 bootcamp per day
    for t_idx in range(n_trainers):
        for d in dates:
            day_vars = [
                x[t_idx, s_idx]
                for s_idx, (d1, d2) in enumerate(slots)
                if (t_idx, s_idx) in x and (d == d1 or d == d2)
            ]
            if day_vars:
                model.add(sum(day_vars) <= 1)

    # Experience: at least 1 experienced trainer per active slot
    if experienced:
        for s_idx in range(n_slots):
            exp_vars = [
                x[t_idx, s_idx]
                for t_idx, name in enumerate(trainer_names)
                if name in experienced and (t_idx, s_idx) in x
            ]
            model.add(sum(exp_vars) >= active[s_idx])

    # Per-trainer weekly cap
    if trainer_info:
        for t_idx, name in enumerate(trainer_names):
            max_pw = (trainer_info.get(name) or {}).get("max_per_week")
            if not max_pw:
                continue
            week_slots: dict[tuple[int, int], list] = defaultdict(list)
            for s_idx, (d1, _) in enumerate(slots):
                if (t_idx, s_idx) in x:
                    yr, wk, _ = d1.isocalendar()
                    week_slots[(yr, wk)].append(x[t_idx, s_idx])
            for week_vars in week_slots.values():
                model.add(sum(week_vars) <= max_pw // 2)

    # Per-week total bootcamp cap
    if weekly_caps:
        week_active: dict[tuple[int, int], list] = defaultdict(list)
        for s_idx, (d1, _) in enumerate(slots):
            yr, wk, _ = d1.isocalendar()
            week_active[(yr, wk)].append(active[s_idx])
        for week_key, week_vars in week_active.items():
            cap = weekly_caps.get(week_key)
            if cap is not None:
                model.add(sum(week_vars) <= cap)

    # Location assignment
    y: dict[tuple[int, int], cp_model.IntVar] = {}
    if locations:
        n_locs = len(locations)
        french_speakers: set[str] = {
            name for name, info in (trainer_info or {}).items()
            if info.get("french_speaking", False)
        }
        for s_idx in range(n_slots):
            for l_idx in range(n_locs):
                y[s_idx, l_idx] = model.new_bool_var(f"y_{s_idx}_{l_idx}")

        # Each active slot assigned to exactly 1 location
        for s_idx in range(n_slots):
            model.add(sum(y[s_idx, l_idx] for l_idx in range(n_locs)) == active[s_idx])

        # French constraint: French-required locations need a French-speaking trainer
        for l_idx, loc in enumerate(locations):
            if loc.get("french_required"):
                for s_idx in range(n_slots):
                    fr_vars = [
                        x[t_idx, s_idx]
                        for t_idx, name in enumerate(trainer_names)
                        if name in french_speakers and (t_idx, s_idx) in x
                    ]
                    model.add(sum(fr_vars) >= y[s_idx, l_idx])

        # Demand cap per location
        for l_idx, loc in enumerate(locations):
            model.add(
                sum(y[s_idx, l_idx] for s_idx in range(n_slots)) <= loc.get("demand", 0)
            )

    BIG_M = 1000
    if weightings:
        weight_bonus = sum(
            weightings.get(name, 0) * x[t_idx, s_idx]
            for t_idx, name in enumerate(trainer_names)
            for s_idx in range(n_slots)
            if (t_idx, s_idx) in x
        )
        model.maximize(BIG_M * sum(active) + weight_bonus)
    else:
        model.maximize(BIG_M * sum(active))

    status = solver.solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        return []

    result = []
    for s_idx, (d1, d2) in enumerate(slots):
        if not solver.value(active[s_idx]):
            continue
        assigned = sorted(
            name for t_idx, name in enumerate(trainer_names)
            if (t_idx, s_idx) in x and solver.value(x[t_idx, s_idx])
        )
        entry: dict = {"slot": (d1, d2), "trainers": assigned}
        if locations:
            for l_idx, loc in enumerate(locations):
                if solver.value(y[s_idx, l_idx]):
                    entry["location"] = loc["name"]
                    break
        result.append(entry)

    return result


def parse_locations(filepath: str) -> list[dict]:
    """Read the Locations sheet from an xlsx file.

    Returns a list of dicts with name, country, demand, french_required.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["Locations"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    locations = []
    for row in rows[1:]:  # skip header
        if not row or row[0] is None:
            continue
        locations.append({
            "name": str(row[0]).strip(),
            "country": str(row[1]).strip() if row[1] else None,
            "demand": int(row[2]) if row[2] is not None else 0,
            "french_required": str(row[3]).strip() == "Yes" if row[3] else False,
        })
    return locations


def apply_bank_holidays(
    trainers: dict[str, dict[date, bool]],
    trainer_info: dict[str, dict],
) -> None:
    """Block trainer availability on public holidays for their home country.

    Modifies trainers in place. Uses the holidays library with a city→country mapping.
    """
    for name, avail in trainers.items():
        info = trainer_info.get(name) or {}
        home = str(info.get("home_location") or "").lower().strip()
        country_code = LOCATION_COUNTRY_CODES.get(home)
        if not country_code:
            continue
        years = {d.year for d in avail}
        country_holidays = hols.country_holidays(country_code, years=years)
        for d in avail:
            if d in country_holidays:
                avail[d] = False


def parse_weightings(filepath: str) -> dict[str, int]:
    """Read the Weightings sheet from an xlsx file.

    Returns a dict mapping trainer name to integer weight.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["Weightings"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    weightings: dict[str, int] = {}
    for row in rows[1:]:  # skip header
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        weightings[name] = int(row[1]) if row[1] is not None else 0
    return weightings
